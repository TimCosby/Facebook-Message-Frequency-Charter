from openpyxl import load_workbook
from openpyxl import Workbook


class maker:

    def __init__(self, reference, start, end):
        self._wb = Workbook()
        self._ws = self._wb.active
        self._reference = reference
        self._max_row = 2
        self._max_col = 1

        self._start_reference_cell = ''
        self._start_reference_index = 0

        # Will be 1 cell further than the last cell last ->end_cell|
        self._end_reference_cell = ''
        self._end_reference_index = 0

        self.set_date(start=start, end=end)

    def set_date(self, start, end):
        col = 2

        while True:
            current_cell = self._reference.cell(row=1, column=col).value

            if current_cell >= start:
                start_index = col
                break
            else:
                col += 1

        new_col = 2
        while True:
            current_cell = self._reference.cell(row=1, column=col).value

            if current_cell > end or current_cell == 'Total':
                # Sets the start and end times for the reference cell
                self._start_reference_cell = self._reference.cell(row=1, column=start_index)
                self._start_reference_index = start_index
                self._end_reference_cell = self._reference.cell(row=1, column=col)
                self._end_reference_index = col

                end_index = col
                # Limit everything within these cells
                self._max_col = end_index - start_index + 2

                self._ws.cell(row=1, column=self._max_col, value='Total')

                break
            else:
                self._ws.cell(row=1, column=new_col, value=current_cell)
                new_col += 1
                col += 1

    def add_person(self, name, row):
        # Add Name
        self._ws.cell(row=self._max_row, column=1, value=name)

        start_cell = 2
        cell_total = 0

        # Add Daily
        for cell in range(self._start_reference_index, self._end_reference_index):
            data = self._reference.cell(row=row, column=cell).value

            if isinstance(data, int):
                cell_total += data
                self._ws.cell(row=self._max_row, column=start_cell, value=data)

            start_cell += 1

        # Add Total
        self._ws.cell(row=self._max_row, column=self._max_col, value=cell_total)
        self._max_row += 1

    def filter(self, name_cache, white_list, blacklist):
        white_filtered = set([])
        black_filtered = set([])

        # White list
        if white_list is not []:
            # If white_list is not empty
            for key in white_list:
                for name in name_cache:
                    # If key is in white list
                    if key.lower() in name.lower():
                        white_filtered.add(name)
        else:
            # If white_list is empty load it with all names
            white_filtered = set(name_cache.keys())

        # Black list
        for key in blacklist:
            for name in name_cache:
                # If key is in black list
                if key.lower() in name.lower():
                    black_filtered.add(name)

        # Remove everything from the white list that is also in the black list
        white_filtered.difference_update(black_filtered)

        # Sort alphabetically
        filtered = list(white_filtered)
        filtered.sort()

        for name in filtered:
            # Add each person
            self.add_person(name, name_cache[name])

    def save(self, name='test'):
        self._wb.save(name + '.xlsx')



def options():
    workbook = load_workbook('database.xlsx')
    worksheet = workbook.active

    name_cache = {}
    removed_names = {}
    i = 2

    for row in range(2, worksheet.max_row + 1):
        # Put all <name>: <row #> in name_cache
        name_cache[worksheet.cell(row=row, column=1).value] = i
        i += 1

    white_list = []
    blacklist = []
    start_date = worksheet.cell(row=1, column=2).value
    end_date = worksheet.cell(row=1, column=worksheet.max_column - 1).value

    while True:
        ans = input('Total Messages for a person? ("total <name>")\n'
                    'Whitelist("add <keyword>")\n'
                    'Blacklist("remove <keyword>"\n'
                    'Within a given time period ("time start <yyyy/mm/dd>" or "time end <yyyy/mm/dd>")? \n'
                    'Get statistics of a certain day ("day <yyyy/mm/dd>">\n'
                    'Done?\n').lower()

        info = ans.split()

        if info is []:
            # If nothing was entered
            pass

        elif 'done' in info[0] or 'exit' in info[0]:
            break

        elif 'add' in info[0]:
            # Adds a keyword to the white_list that has to appear in a person's name
            white_list.append(info[1])

        elif 'remove' in info[0]:
            # Adds a keyword to the blacklist that has to appear in a person's name
            blacklist.append(info[1])

        elif 'time' == info[0]:
            # Limit the time in the graph
            try:
                if info[2][:4].isnumeric() and info[2][4] == info[2][7] == '/' and info[2][5:7].isnumeric() and info[2][8:].isnumeric():
                    if 'start' == info[1]:
                        if info[2] < end_date:
                            start_date = info[2]
                        else:
                            print('Start date exceeds the end date!')
                    elif 'end' == info[1]:
                        if info[2] > start_date:
                            end_date = info[2]
                        else:
                            print('End date exceeds the start date!')
            except: print('Invalid Date!')

        elif 'total' in info[0]:
            # Get total reiceved msgs from a person
            try:
                attempted_people = []

                for name in name_cache:
                    # Tests all names in database
                    if info[1] in name.lower():
                        # If name matches the keyword
                        attempted_people.append(name)

                if len(attempted_people) == 0:
                    # If no results match
                    print('Person does not exist!')
                elif len(attempted_people) == 1:
                    # If only one result appears
                    print('You have a total of: ' + str(worksheet.cell(row=name_cache[attempted_people[0]], column=worksheet.max_column).value) + ' messages from ' + attempted_people[0] + '\n')
                else:
                    # If multiple results appear
                    print('Did you mean:')

                    for i in range(len(attempted_people)):
                        # Post all the results of people that showed up and asks user to pick one
                        print(str(i + 1) + '. ' + attempted_people[i] + '\n')

                    ans = input('Enter the number of the person you want to search, otherwise enter any other key\n')
                    if ans.isnumeric() and int(ans) - 1 <= len(attempted_people):
                        print('You have a total of: ' + str(worksheet.cell(row=name_cache[attempted_people[int(ans) - 1]], column=worksheet.max_column).value) + ' messages from ' + attempted_people[int(ans) - 1] + '\n')
            except:
                # If invalid information was entered
                print('Person does not exist!\n')

    ws = maker(worksheet, start_date, end_date)

    if white_list is not [] or blacklist is not []:
        ws.filter(name_cache, white_list, blacklist)
    else:
        # Add everyone
        name_list = list(name_cache.keys())
        name_list.sort()
        for name in name_list:
            ws.add_person(name, name_cache[name])

    ws.save()
