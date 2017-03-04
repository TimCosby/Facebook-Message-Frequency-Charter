from bs4 import BeautifulSoup
from facebook import GraphAPI
import threading
from queue import Queue
from openpyxl import Workbook

TOKEN = 'EAAFpZCp8376ABANHxBKhCyYgfBZAZBi6J41eE4UBJL2fkIZCRUFVI3KB0wM3ZAlli2KDi72bxYkTX89t2FNUiXKmQYbSDaBRnxg3Upmgh1hiJNZBnT2cby6tBg501RHDYh3uWZCCichnFGiATf7fJZAiCDAHKuKLTYWWsmwYOvukUQZDZD'
DAYS = ['Monday, ', 'Tuesday, ', 'Wednesday, ', 'Thursday, ', 'Friday, ', 'Saturday, ', 'Sunday, ']
MONTHS = {'January': '01', 'February': '02', 'March': '03', 'April': '04', 'May': '05', 'June': '06', "July": '07', 'August': '08', 'September': '09', 'October': '10', 'November': '11', 'December': '12'}
DATE_ADJUST = {'1': '01', '2': '02', '3': '03', '4': '04', '5': '05', '6': '06', '7': '07', '8': '08', '9': '09'}
CACHE_ID = {}

TOTAL = [0]
COUNT = [0]


def get_name(id):
    """
    Get id's Facebook Name

    :param int id: Facebook ID
    :return: str
    """

    graph = GraphAPI(access_token=TOKEN, version='2.5')

    return graph.get_object(id=str(id).split('-')[0])['name']


def add_option(message_count, que, threads):
    """
    Find date and name of user then add to dictionary

    :param dict message_count:
    :param Queue que:
    :return: None
    """

    print(str(COUNT[0]) + ' / ' + str(TOTAL[0]))

    COUNT[0] += 1

    clean_line = que.get()

    for day in DAYS:
        if day in clean_line[:clean_line.find('at')]:
            line_start = clean_line.find(day) + len(day)
            line_end = clean_line.find(' at')

            date = clean_line[line_start:line_end]

            date = date.split(' ')
            # Remove comma
            date[1] = date[1][:-1]
            try:
                date[1] = DATE_ADJUST[date[1]]
            except: pass
            # Turn into numeric value
            date[0] = MONTHS[date[0]]

            if clean_line[0].isnumeric():
                numeric_option(message_count, date, int(clean_line[:line_start - len(day)].split('@')[-2]))
            elif clean_line[0].isalpha():
                alpha_option(message_count, date, clean_line[:line_start - len(day)])

    try:
        threads.remove(threading.current_thread())
    except:pass

    que.task_done()


def numeric_option(message_count, date, id):
    """
    Add one to the dictionary of format {date}{name} = <Number of Messages>

    :param dict message_count:
    :param list of str date:
    :param int id:
    :return: None
    """

    if id in CACHE_ID and CACHE_ID[id] != None:
        alpha_option(message_count, date, CACHE_ID[id])
    else:
        try:
            CACHE_ID[id] = get_name(id)
            alpha_option(message_count, date, CACHE_ID[id])
        except:
            CACHE_ID[id] = None


def alpha_option(message_count, date, name):
    """
    TO-DO
    Put in {Yr:{Mon:{Day}}}

    :param message_count:
    :param list of str date:
    :param name:
    :return:
    """

    if date[2] in message_count:
        # If year is already cached
        if date[0] in message_count[date[2]]:
            # If month is already cached
            if date[1] in message_count[date[2]][date[0]]:
                # If day is already cached
                if name in message_count[date[2]][date[0]][date[1]]:
                    # Add one to the message cont
                    message_count[date[2]][date[0]][date[1]][name] += 1
                else:
                    # If user has not yet sent a message that day
                    message_count[date[2]][date[0]][date[1]][name] = 1
            else:
                # If day is not cached
                message_count[date[2]][date[0]][date[1]] = {}
                # Recurse
                alpha_option(message_count, date, name)
        else:
            # If month is not cached
            message_count[date[2]][date[0]] = {}
            message_count[date[2]][date[0]][date[1]] = {}
            # Recurse
            alpha_option(message_count, date, name)
    else:
        # If year is not cached
        message_count[date[2]] = {}
        message_count[date[2]][date[0]] = {}
        message_count[date[2]][date[0]][date[1]] = {}
        # Recurse
        alpha_option(message_count, date, name)


def get_messages(message_count):
    """
    Get the message count for each user you message

    :param dict message_count:
    :return: NoneType
    """

    file = open('messages.htm', encoding='UTF-8')

    html = file.read().split('</p>')
    file.close()

    TOTAL[0] = len(html) - 1

    # Gets rid of formatting at the beginning
    start = html[0].find('<div class="message">')
    while not html[0][start].isnumeric():
        start += 1
    html[0] = html[0][start:]

    html.pop()

    threads = []

    que = Queue(maxsize=50)
    for line in html:
        clean_line = BeautifulSoup(line, 'lxml').getText()
        #print(line)
        if len(clean_line) != 0:
            t = threading.Thread(target=add_option,
                                 args=(message_count, que, threads))
            que.put(clean_line)

            t.daemon = True
            t.start()
            threads.append(t)

    que.join()


def total_of_row(worksheet, row):
    """
    Return total values in row <row>

    :param worksheet:
    :param row:
    :return:
    """

    total = 0

    for (cell,) in worksheet.iter_cols(min_col=2, min_row=row, max_col=worksheet.max_column, max_row=row):
        if cell.value != None:
            total += cell.value

    return total


def create_file(dict):
    """
    Records dates, names, and messages of name per date

    :return: None
    :output: Excel File
    """

    workbook = Workbook()
    worksheet = workbook.active

    name_cache = {}
    col = 2

    # Sorts all the years in dictionary from past -> present
    sorted_years = list(dict.keys())

    sorted_years.sort()

    for year in sorted_years:
        # Sorts all the months in dictionary[year] from past -> present
        sorted_months = list(dict[year].keys())

        for month in range(len(sorted_months)):
            # If a single digit add a 0 in front of it so the sort works properly
            sorted_months[month].rjust(2, '0')

        sorted_months.sort()

        for month in sorted_months:
            # Sorts all the days in dictionary[year][month] from past -> present
            sorted_days = list(dict[year][month].keys())

            for day in range(len(sorted_days)):
                # If a single digit add a 0 in front of it so the sort works properly
                sorted_days[day] = sorted_days[day].rjust(2, '0')

            sorted_days.sort()

            for day in sorted_days:
                # Records a new date
                worksheet.cell(row=1, column=col, value=year + '/' + month + '/' + day)

                # Records a person
                for name in dict[year][month][day]:
                    try:
                        name_cache[name]
                    except:
                        name_cache[name] = len(name_cache) + 2

                    worksheet.cell(row=name_cache[name], column=1, value=name)

                    worksheet.cell(row=name_cache[name], column=col, value=dict[year][month][day][name])

                col += 1


    # Get total msgs/person
    max_column = worksheet.max_column
    worksheet.cell(row=1, column=max_column + 1, value='Total')

    for row in range(2, worksheet.max_row + 1):
        total = total_of_row(worksheet, row)

        worksheet.cell(row=row, column=max_column + 1, value=total)

    # Try to save
    try:
        workbook.save('database.xlsx')
    except:
        workbook.save('database1.xlsx')
